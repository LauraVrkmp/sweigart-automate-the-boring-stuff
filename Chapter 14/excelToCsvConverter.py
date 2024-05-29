import openpyxl, csv, os
from openpyxl.utils.cell import get_column_letter

for file in os.listdir('excelSpreadsheets\\'):
    if file.endswith('.xlsx'):
        wb = openpyxl.load_workbook('excelSpreadsheets\\' + file)
        for sheet in wb.sheetnames:
            current_sheet = wb[sheet]
            csvFilename = str(file)[:-5] + '_' + str(sheet) + '.csv'
            csvFile = open(os.path.join('excelSpreadsheets', csvFilename), 
                           'w', newline='')
            
            csvWriter = csv.writer(csvFile)
            for rowNum in range(1, current_sheet.max_row + 1):
                rowData = []
                for colNum in range(1, current_sheet.max_column + 1):
                    rowData.append(current_sheet[
                        get_column_letter(colNum) + str(rowNum)].value)
                csvWriter.writerow(rowData)

            csvFile.close()