import openpyxl, sys
from openpyxl.utils.cell import get_column_letter

N = int(sys.argv[1])
M = int(sys.argv[2])
file = sys.argv[3]

wb = openpyxl.load_workbook(file)
sheet = wb['Population by Census Tract']

new_wb = openpyxl.Workbook()
new_wb.create_sheet(title='Population by Census Tract')
new_sheet = new_wb['Population by Census Tract']
new_wb.remove(new_wb['Sheet'])

for i in range(1, sheet.max_row + 1):
    if i == N:
        for j in range(1, sheet.max_column + 1):
            new_sheet[get_column_letter(j) + str(i)].value = sheet[get_column_letter(j) + str(i)].value
        for skip in range(1, M + 1):
            for j in range(1, sheet.max_column + 1):
                new_sheet[get_column_letter(j) + str(i + skip)].value = ''
        i += M
    for j in range(1, sheet.max_column + 1):
        new_sheet[get_column_letter(j) + str(i + (i > N) * M)].value = sheet[get_column_letter(j) + str(i)].value

new_wb.save('inserted_' + file)