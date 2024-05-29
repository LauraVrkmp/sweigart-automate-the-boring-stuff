import openpyxl, os
from pathlib import Path
from openpyxl.utils.cell import get_column_letter

wb = openpyxl.load_workbook('text_to_spreadsheet.xlsx')
sheet = wb['Sheet']

path = Path('Text files')
if not path.exists():
    path.mkdir()

for i in range(1, sheet.max_column + 1):
    file = open('Text files\\file' + str(i) + '.txt', 'w')
    for j in range(1, sheet.max_row + 1):
        line = sheet[get_column_letter(i) + str(j)].value
        file.write(str(line))
    file.close()