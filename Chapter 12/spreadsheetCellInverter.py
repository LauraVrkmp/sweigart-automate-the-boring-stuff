import openpyxl
from openpyxl.utils.cell import get_column_letter

wb = openpyxl.load_workbook('censuspopdata_shortened.xlsx')
sheet = wb['Population by Census Tract']

new_wb = openpyxl.Workbook()
new_wb.create_sheet(title='Population by Census Tract')
new_sheet = new_wb['Population by Census Tract']
new_wb.remove(new_wb['Sheet'])

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        new_sheet[get_column_letter(i) + str(j)].value = sheet[get_column_letter(j) + str(i)].value

new_wb.save('inverted_censuspopdata_shortened.xlsx')