import openpyxl, sys
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb['Sheet']

N = int(sys.argv[1])

# Fill the first column
for i in range(1, N + 1):
    sheet['A' + str(i+1)].font = Font(bold=True)
    sheet['A' + str(i+1)] = i

# Fill the first row
for i in range(1, N + 1):
    sheet[get_column_letter(i + 1) + '1'].font = Font(bold=True)
    sheet[get_column_letter(i + 1) + '1'] = i

# Fill the multiplication cells
for i in range(1, N + 1):
    for j in range(1, N + 1):
        sheet[get_column_letter(i + 1) + str(j + 1)] = \
        '=' + 'A' + str(i+1) + '*' + get_column_letter(j+1) + '1'

wb.save('multiplicationTable.xlsx')


# get_column_letter()