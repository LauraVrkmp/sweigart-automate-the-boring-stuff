import openpyxl, psutil
from openpyxl.utils.cell import get_column_letter

wb = openpyxl.Workbook()
sheet = wb['Sheet']

file1 = open('Text files/file1.txt')
lines1 = file1.readlines()

file2 = open('Text files/file2.txt')
lines2 = file2.readlines()

file3 = open('Text files/file3.txt')
lines3 = file3.readlines()

proc = psutil.Process()

for i in range(1, len(proc.open_files())):
    lines = eval('lines' + str(i))
    for j in range(len(lines)):
        sheet[get_column_letter(i) + str(j + 1)].value = lines[j]

wb.save('text_to_spreadsheet.xlsx')

file1.close()
file2.close()
file3.close()